"""엑셀의 테스트 문장 100개를 LLM에 일괄 입력하고 정확도를 계산한다.

이 파일을 프로젝트의 src 폴더에 넣고 실행하는 것을 기준으로 한다.

사용 예:
    python src/batch_accuracy_test.py "예제 100.xlsm"
    python src/batch_accuracy_test.py "예제 100.xlsm" --output "테스트 결과.xlsm"
    python src/batch_accuracy_test.py "예제 100.xlsm" --resume
"""

from __future__ import annotations

import argparse
import json
import math
import time
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from agent import json_to_command, process


SHEET_NAME = "예시 문장"
SUMMARY_SHEET_NAME = "정확도 요약"

# 엑셀 열 번호
COL_NUMBER = 1
COL_PROCESS = 2
COL_MATERIAL = 3
COL_X = 4
COL_Y = 5
COL_Z = 6
COL_RPM = 7
COL_DEPTH = 8
COL_TOOL = 9
COL_DIAMETER = 10
COL_SENTENCE = 11
COL_RAW_JSON = 12
COL_RESULT = 13
COL_NOTE = 14

FIELDS = (
    "process",
    "material",
    "x",
    "y",
    "z",
    "rpm",
    "depth",
    "tool",
    "diameter",
)

EXPECTED_COLUMNS = {
    "process": COL_PROCESS,
    "material": COL_MATERIAL,
    "x": COL_X,
    "y": COL_Y,
    "z": COL_Z,
    "rpm": COL_RPM,
    "depth": COL_DEPTH,
    "tool": COL_TOOL,
    "diameter": COL_DIAMETER,
}

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
WHITE_BOLD_FONT = Font(color="FFFFFF", bold=True)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="엑셀 테스트 문장을 일괄 실행하고 LLM 명령 해석 정확도를 계산합니다."
    )
    parser.add_argument("workbook", type=Path, help="입력 .xlsx 또는 .xlsm 파일")
    parser.add_argument(
        "--output",
        type=Path,
        help="결과 파일 경로(생략하면 입력 파일명 뒤에 _테스트결과를 붙임)",
    )
    parser.add_argument(
        "--start-row",
        type=int,
        default=2,
        help="테스트 시작 행(기본값: 2)",
    )
    parser.add_argument(
        "--end-row",
        type=int,
        default=101,
        help="테스트 종료 행(기본값: 101)",
    )
    parser.add_argument(
        "--delay",
        type=float,
        default=0.2,
        help="API 호출 사이 대기 시간(초, 기본값: 0.2)",
    )
    parser.add_argument(
        "--resume",
        action="store_true",
        help="M열이 정답/오답/오류인 행은 다시 호출하지 않음",
    )
    parser.add_argument(
        "--max-retries",
        type=int,
        default=3,
        help="API 호출 실패 시 최대 시도 횟수(기본값: 3)",
    )
    return parser.parse_args()


def default_output_path(input_path: Path) -> Path:
    return input_path.with_name(
        f"{input_path.stem}_테스트결과{input_path.suffix}"
    )


def call_process_with_retry(
    sentence: str,
    max_retries: int,
) -> tuple[list[dict[str, Any]], str]:
    last_error: Exception | None = None

    for attempt in range(1, max_retries + 1):
        try:
            return process(sentence)
        except Exception as error:  # API·네트워크 오류를 행 단위로 기록하기 위함
            last_error = error
            if attempt < max_retries:
                wait_seconds = 2 ** (attempt - 1)
                print(
                    f"  호출 실패({attempt}/{max_retries}): {error}\n"
                    f"  {wait_seconds}초 후 다시 시도합니다."
                )
                time.sleep(wait_seconds)

    assert last_error is not None
    raise last_error


def commands_to_actual(
    parsed_commands: list[dict[str, Any]],
) -> tuple[dict[str, Any], list[str]]:
    actual = {field: None for field in FIELDS}
    errors: list[str] = []

    for parsed in parsed_commands:
        if not parsed.get("valid"):
            errors.append(str(parsed.get("error", "알 수 없는 명령 오류")))
            continue

        command = parsed.get("command")
        if command is None:
            errors.append("Command 객체가 없습니다.")
            continue

        if command.name == "target":
            position = command.args.get("position") or {}
            for field in ("x", "y", "z"):
                if position.get(field) is not None:
                    actual[field] = position[field]

        elif command.name == "machine":
            for field in (
                "process",
                "material",
                "rpm",
                "depth",
                "tool",
                "diameter",
            ):
                if command.args.get(field) is not None:
                    actual[field] = command.args[field]

    return actual, errors


def expected_from_row(sheet, row: int) -> dict[str, Any]:
    return {
        field: sheet.cell(row=row, column=column).value
        for field, column in EXPECTED_COLUMNS.items()
    }


def values_equal(expected: Any, actual: Any) -> bool:
    if expected is None or actual is None:
        return expected is actual

    if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
        return math.isclose(
            float(expected),
            float(actual),
            rel_tol=1e-9,
            abs_tol=1e-6,
        )

    return str(expected).strip().casefold() == str(actual).strip().casefold()


def compare_fields(
    expected: dict[str, Any],
    actual: dict[str, Any],
) -> tuple[list[str], dict[str, bool]]:
    field_results: dict[str, bool] = {}
    mismatch_messages: list[str] = []

    for field in FIELDS:
        is_correct = values_equal(expected[field], actual[field])
        field_results[field] = is_correct

        if not is_correct:
            mismatch_messages.append(
                f"{field}: 정답={expected[field]!r}, 실제={actual[field]!r}"
            )

    return mismatch_messages, field_results


def raw_output_as_single_line(raw_output: str) -> str:
    try:
        return json.dumps(
            json.loads(raw_output),
            ensure_ascii=False,
            separators=(",", ":"),
        )
    except (json.JSONDecodeError, TypeError):
        return raw_output


def parsed_commands_from_raw(raw_output: str) -> list[dict[str, Any]]:
    data = json.loads(raw_output)
    if isinstance(data, dict):
        data = [data]
    if not isinstance(data, list):
        raise ValueError("저장된 LLM 출력이 JSON 배열 또는 객체가 아닙니다.")
    return [
        json_to_command(item)
        for item in data
        if isinstance(item, dict)
    ]


def prepare_result_columns(sheet) -> None:
    sheet.cell(1, COL_RAW_JSON).value = "실제 LLM JSON"
    sheet.cell(1, COL_RESULT).value = "결과"
    sheet.cell(1, COL_NOTE).value = "비고"

    for column in (COL_RAW_JSON, COL_RESULT, COL_NOTE):
        cell = sheet.cell(1, column)
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    sheet.column_dimensions["L"].width = 55
    sheet.column_dimensions["M"].width = 12
    sheet.column_dimensions["N"].width = 65


def write_summary(
    workbook,
    rows_tested: int,
    sentence_correct: int,
    field_correct: Counter,
    field_total: Counter,
    error_count: int,
) -> None:
    if SUMMARY_SHEET_NAME in workbook.sheetnames:
        del workbook[SUMMARY_SHEET_NAME]

    summary = workbook.create_sheet(SUMMARY_SHEET_NAME, 1)
    summary.sheet_view.showGridLines = False

    sentence_accuracy = (
        sentence_correct / rows_tested if rows_tested else 0
    )
    total_correct = sum(field_correct.values())
    total_fields = sum(field_total.values())
    field_accuracy = total_correct / total_fields if total_fields else 0

    summary["A1"] = "LLM 명령 해석 정확도"
    summary["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    summary["A1"].fill = HEADER_FILL
    summary.merge_cells("A1:C1")
    summary["A1"].alignment = Alignment(horizontal="center")

    summary.append(["지표", "결과", "설명"])
    summary.append(["테스트 문장 수", rows_tested, "실제로 평가된 행 수"])
    summary.append(["완전 정답 문장 수", sentence_correct, "9개 필드가 모두 일치"])
    summary.append(["문장 단위 정확도", sentence_accuracy, "완전 일치 기준"])
    summary.append(["필드 단위 정확도", field_accuracy, "전체 필드 일치 비율"])
    summary.append(["실행 오류 수", error_count, "API 또는 명령 변환 오류"])

    for cell in summary[2]:
        cell.fill = SUBHEADER_FILL
        cell.font = Font(bold=True)

    start_row = 9
    summary.cell(start_row, 1).value = "필드"
    summary.cell(start_row, 2).value = "정답 수"
    summary.cell(start_row, 3).value = "정확도"

    for cell in summary[start_row]:
        cell.fill = SUBHEADER_FILL
        cell.font = Font(bold=True)

    for offset, field in enumerate(FIELDS, start=1):
        row = start_row + offset
        total = field_total[field]
        summary.cell(row, 1).value = field
        summary.cell(row, 2).value = field_correct[field]
        summary.cell(row, 3).value = (
            field_correct[field] / total if total else 0
        )

    summary["B5"].number_format = "0.0%"
    summary["B6"].number_format = "0.0%"
    summary.cell(start_row + 1, 3).number_format = "0.0%"
    for row in range(start_row + 1, start_row + len(FIELDS) + 1):
        summary.cell(row, 3).number_format = "0.0%"

    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 16
    summary.column_dimensions["C"].width = 36
    summary.freeze_panes = "A3"


def main() -> None:
    args = parse_args()
    input_path = args.workbook.resolve()
    output_path = (
        args.output.resolve()
        if args.output is not None
        else default_output_path(input_path)
    )

    if not input_path.exists():
        raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {input_path}")

    keep_vba = input_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(input_path, keep_vba=keep_vba)

    if SHEET_NAME not in workbook.sheetnames:
        raise KeyError(
            f"'{SHEET_NAME}' 시트를 찾을 수 없습니다. "
            f"현재 시트: {', '.join(workbook.sheetnames)}"
        )

    sheet = workbook[SHEET_NAME]
    prepare_result_columns(sheet)

    sentence_correct = 0
    field_correct: Counter = Counter()
    field_total: Counter = Counter()
    error_count = 0
    rows_tested = 0

    end_row = min(args.end_row, sheet.max_row)

    for row in range(args.start_row, end_row + 1):
        sentence = sheet.cell(row, COL_SENTENCE).value
        previous_result = sheet.cell(row, COL_RESULT).value

        if not sentence:
            continue

        test_number = sheet.cell(row, COL_NUMBER).value
        expected = expected_from_row(sheet, row)

        if args.resume and previous_result in {"정답", "오답", "오류"}:
            print(f"[{test_number}/100] 기존 LLM 결과를 다시 비교합니다.")
            raw_output = sheet.cell(row, COL_RAW_JSON).value or ""
            try:
                parsed_commands = parsed_commands_from_raw(str(raw_output))
                actual, command_errors = commands_to_actual(parsed_commands)
                mismatches, field_results = compare_fields(expected, actual)

                rows_tested += 1
                for field, is_correct in field_results.items():
                    field_total[field] += 1
                    if is_correct:
                        field_correct[field] += 1

                if command_errors:
                    error_count += 1
                elif not mismatches:
                    sentence_correct += 1
            except Exception:
                rows_tested += 1
                error_count += 1
                for field in FIELDS:
                    field_total[field] += 1
            continue

        print(f"[{test_number}/100] {sentence}")

        try:
            parsed_commands, raw_output = call_process_with_retry(
                str(sentence),
                args.max_retries,
            )
            actual, command_errors = commands_to_actual(parsed_commands)
            mismatches, field_results = compare_fields(expected, actual)

            rows_tested += 1
            for field, is_correct in field_results.items():
                field_total[field] += 1
                if is_correct:
                    field_correct[field] += 1

            sheet.cell(row, COL_RAW_JSON).value = raw_output_as_single_line(
                raw_output
            )

            notes = command_errors + mismatches
            if command_errors:
                error_count += 1
                sheet.cell(row, COL_RESULT).value = "오류"
            elif mismatches:
                sheet.cell(row, COL_RESULT).value = "오답"
            else:
                sentence_correct += 1
                sheet.cell(row, COL_RESULT).value = "정답"

            sheet.cell(row, COL_NOTE).value = "\n".join(notes)

        except Exception as error:
            rows_tested += 1
            error_count += 1
            for field in FIELDS:
                field_total[field] += 1

            sheet.cell(row, COL_RAW_JSON).value = ""
            sheet.cell(row, COL_RESULT).value = "오류"
            sheet.cell(row, COL_NOTE).value = (
                f"{type(error).__name__}: {error}"
            )

        for column in (COL_RAW_JSON, COL_NOTE):
            sheet.cell(row, column).alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )
        sheet.cell(row, COL_RESULT).alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        # 중간에 종료되더라도 결과를 최대한 보존한다.
        if rows_tested % 5 == 0:
            write_summary(
                workbook,
                rows_tested,
                sentence_correct,
                field_correct,
                field_total,
                error_count,
            )
            workbook.save(output_path)
            print(f"  중간 저장: {output_path}")

        if args.delay > 0:
            time.sleep(args.delay)

    write_summary(
        workbook,
        rows_tested,
        sentence_correct,
        field_correct,
        field_total,
        error_count,
    )
    workbook.save(output_path)

    sentence_accuracy = (
        sentence_correct / rows_tested * 100 if rows_tested else 0
    )
    total_fields = sum(field_total.values())
    total_correct = sum(field_correct.values())
    field_accuracy = (
        total_correct / total_fields * 100 if total_fields else 0
    )

    print("\n테스트 완료")
    print(f"결과 파일: {output_path}")
    print(f"문장 단위 정확도: {sentence_accuracy:.1f}%")
    print(f"필드 단위 정확도: {field_accuracy:.1f}%")
    print(f"실행 오류: {error_count}개")


if __name__ == "__main__":
    main()
