"""녹음 파일 100개의 Whisper·정규화·최종 명령 정확도를 평가한다.

이 파일을 프로젝트의 src 폴더에 넣고 실행한다.

사용 예:
    python src/evaluate_voice_accuracy.py "예제 100.xlsm"
    python src/evaluate_voice_accuracy.py "예제 100.xlsm" --skip-llm
    python src/evaluate_voice_accuracy.py "예제 100.xlsm" --resume
"""

from __future__ import annotations

import argparse
import math
import re
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Any, Sequence

import numpy as np
import scipy.io.wavfile as wav
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from agent import process
from voice.normalizer import normalize_voice_command
from voice.recognizer import transcribe_audio
from voice.recorder import SAMPLE_RATE


SHEET_NAME = "예시 문장"
SUMMARY_SHEET = "음성 정확도 요약"

COL_NUMBER = 1
COL_SENTENCE = 11
COL_AUDIO = 15
COL_WHISPER = 16
COL_NORMALIZED = 17
COL_CER = 18
COL_WER = 19
COL_RAW_EXACT = 20
COL_NORMALIZED_EXACT = 21
COL_LLM_JSON = 22
COL_COMMAND_RESULT = 23
COL_NOTE = 24

FIELDS = (
    "process",
    "material",
    "x",
    "y",
    "z",
    "rpm",
    "depth",
    "tool",
    "diameter",
)

EXPECTED_COLUMNS = {
    "process": 2,
    "material": 3,
    "x": 4,
    "y": 5,
    "z": 6,
    "rpm": 7,
    "depth": 8,
    "tool": 9,
    "diameter": 10,
}

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
WHITE_BOLD_FONT = Font(color="FFFFFF", bold=True)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Whisper 인식률과 최종 명령 정확도를 자동 평가합니다."
    )
    parser.add_argument("workbook", type=Path, help="정답 엑셀 파일")
    parser.add_argument(
        "--audio-dir",
        type=Path,
        default=Path("voice_test_audio"),
        help="001.wav~100.wav 폴더",
    )
    parser.add_argument("--output", type=Path, help="결과 엑셀 경로")
    parser.add_argument(
        "--skip-llm",
        action="store_true",
        help="Whisper와 정규화만 평가하고 LLM 호출은 생략",
    )
    parser.add_argument(
        "--resume",
        action="store_true",
        help="P열에 기존 Whisper 결과가 있는 행은 다시 인식하지 않음",
    )
    parser.add_argument("--start", type=int, default=1, help="시작 번호")
    parser.add_argument("--end", type=int, default=100, help="종료 번호")
    return parser.parse_args()


def default_output_path(input_path: Path) -> Path:
    return input_path.with_name(
        f"{input_path.stem}_음성테스트결과{input_path.suffix}"
    )


def edit_distance(
    reference: Sequence[str],
    hypothesis: Sequence[str],
) -> int:
    previous = list(range(len(hypothesis) + 1))

    for ref_index, ref_item in enumerate(reference, start=1):
        current = [ref_index]
        for hyp_index, hyp_item in enumerate(hypothesis, start=1):
            insertion = current[hyp_index - 1] + 1
            deletion = previous[hyp_index] + 1
            substitution = previous[hyp_index - 1] + (
                0 if ref_item == hyp_item else 1
            )
            current.append(min(insertion, deletion, substitution))
        previous = current

    return previous[-1]


def clean_for_cer(text: str) -> list[str]:
    text = unicodedata.normalize("NFKC", text).casefold()
    text = re.sub(r"[\s\W_]+", "", text, flags=re.UNICODE)
    return list(text)


def clean_for_wer(text: str) -> list[str]:
    text = unicodedata.normalize("NFKC", text).casefold()
    text = re.sub(r"[^\w]+", " ", text, flags=re.UNICODE)
    return text.split()


def error_rate(reference: Sequence[str], hypothesis: Sequence[str]) -> float:
    if not reference:
        return 0.0 if not hypothesis else 1.0
    return edit_distance(reference, hypothesis) / len(reference)


def text_metrics(reference: str, hypothesis: str) -> tuple[float, float]:
    return (
        error_rate(clean_for_cer(reference), clean_for_cer(hypothesis)),
        error_rate(clean_for_wer(reference), clean_for_wer(hypothesis)),
    )


def read_audio(path: Path) -> np.ndarray:
    sample_rate, audio = wav.read(path)

    if sample_rate != SAMPLE_RATE:
        raise ValueError(
            f"샘플레이트가 {sample_rate}Hz입니다. "
            f"{SAMPLE_RATE}Hz WAV 파일이 필요합니다."
        )

    if audio.ndim == 2:
        if audio.shape[1] == 1:
            audio = audio[:, 0]
        else:
            audio = np.mean(audio.astype(np.float32), axis=1)

    if audio.dtype != np.int16:
        if np.issubdtype(audio.dtype, np.floating):
            audio = np.clip(audio, -1.0, 1.0)
            audio = (audio * 32767).astype(np.int16)
        else:
            audio = audio.astype(np.int16)

    return audio


def expected_from_row(sheet, row: int) -> dict[str, Any]:
    return {
        field: sheet.cell(row, column).value
        for field, column in EXPECTED_COLUMNS.items()
    }


def commands_to_actual(
    parsed_commands: list[dict[str, Any]],
) -> tuple[dict[str, Any], list[str]]:
    actual = {field: None for field in FIELDS}
    errors: list[str] = []

    for parsed in parsed_commands:
        if not parsed.get("valid"):
            errors.append(str(parsed.get("error", "명령 변환 오류")))
            continue

        command = parsed.get("command")
        if command is None:
            errors.append("Command 객체가 없습니다.")
            continue

        if command.name == "target":
            position = command.args.get("position") or {}
            for field in ("x", "y", "z"):
                if position.get(field) is not None:
                    actual[field] = position[field]

        elif command.name == "machine":
            for field in (
                "process",
                "material",
                "rpm",
                "depth",
                "tool",
                "diameter",
            ):
                if command.args.get(field) is not None:
                    actual[field] = command.args[field]

    return actual, errors


def values_equal(expected: Any, actual: Any) -> bool:
    if expected is None or actual is None:
        return expected is actual
    if isinstance(expected, (int, float)) and isinstance(actual, (int, float)):
        return math.isclose(
            float(expected),
            float(actual),
            rel_tol=1e-9,
            abs_tol=1e-6,
        )
    return str(expected).strip().casefold() == str(actual).strip().casefold()


def compare_commands(
    expected: dict[str, Any],
    actual: dict[str, Any],
) -> tuple[list[str], dict[str, bool]]:
    field_results: dict[str, bool] = {}
    mismatches: list[str] = []

    for field in FIELDS:
        correct = values_equal(expected[field], actual[field])
        field_results[field] = correct
        if not correct:
            mismatches.append(
                f"{field}: 정답={expected[field]!r}, 실제={actual[field]!r}"
            )

    return mismatches, field_results


def prepare_columns(sheet) -> None:
    headers = {
        COL_AUDIO: "음성 파일",
        COL_WHISPER: "Whisper 원문",
        COL_NORMALIZED: "정규화 결과",
        COL_CER: "CER",
        COL_WER: "WER",
        COL_RAW_EXACT: "Whisper 완전 일치",
        COL_NORMALIZED_EXACT: "정규화 후 완전 일치",
        COL_LLM_JSON: "음성 기반 LLM JSON",
        COL_COMMAND_RESULT: "최종 명령 결과",
        COL_NOTE: "음성 테스트 비고",
    }

    for column, title in headers.items():
        cell = sheet.cell(1, column)
        cell.value = title
        cell.fill = HEADER_FILL
        cell.font = WHITE_BOLD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for letter, width in {
        "O": 14,
        "P": 55,
        "Q": 55,
        "R": 10,
        "S": 10,
        "T": 16,
        "U": 18,
        "V": 55,
        "W": 16,
        "X": 65,
    }.items():
        sheet.column_dimensions[letter].width = width


def write_summary(
    workbook,
    tested: int,
    raw_exact: int,
    normalized_exact: int,
    cer_sum: float,
    wer_sum: float,
    command_correct: int,
    command_tested: int,
    field_correct: Counter,
    field_total: Counter,
    missing_audio: int,
) -> None:
    if SUMMARY_SHEET in workbook.sheetnames:
        del workbook[SUMMARY_SHEET]

    sheet = workbook.create_sheet(SUMMARY_SHEET, 1)
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:C1")
    sheet["A1"] = "음성인식 및 최종 명령 정확도"
    sheet["A1"].fill = HEADER_FILL
    sheet["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    sheet["A1"].alignment = Alignment(horizontal="center")

    sheet.append(["지표", "결과", "설명"])
    rows = [
        ("평가 음성 수", tested, "실제로 Whisper 평가한 WAV 수"),
        ("누락 음성 수", missing_audio, "WAV 파일이 없는 문장 수"),
        ("평균 CER", cer_sum / tested if tested else 0, "낮을수록 좋음"),
        (
            "문자 정확도",
            max(0, 1 - cer_sum / tested) if tested else 0,
            "1 - 평균 CER",
        ),
        ("평균 WER", wer_sum / tested if tested else 0, "낮을수록 좋음"),
        (
            "단어 정확도",
            max(0, 1 - wer_sum / tested) if tested else 0,
            "1 - 평균 WER",
        ),
        ("Whisper 완전 일치율", raw_exact / tested if tested else 0, "원문 기준"),
        (
            "정규화 후 완전 일치율",
            normalized_exact / tested if tested else 0,
            "normalizer 적용 후",
        ),
        (
            "최종 명령 정확도",
            command_correct / command_tested if command_tested else 0,
            "9개 필드 완전 일치",
        ),
    ]

    for row in rows:
        sheet.append(row)

    for cell in sheet[2]:
        cell.fill = SUBHEADER_FILL
        cell.font = Font(bold=True)

    for row in range(5, 12):
        sheet.cell(row, 2).number_format = "0.0%"

    start_row = 14
    sheet.cell(start_row, 1).value = "필드"
    sheet.cell(start_row, 2).value = "정답 수"
    sheet.cell(start_row, 3).value = "정확도"
    for cell in sheet[start_row]:
        cell.fill = SUBHEADER_FILL
        cell.font = Font(bold=True)

    for offset, field in enumerate(FIELDS, start=1):
        row = start_row + offset
        total = field_total[field]
        sheet.cell(row, 1).value = field
        sheet.cell(row, 2).value = field_correct[field]
        sheet.cell(row, 3).value = (
            field_correct[field] / total if total else 0
        )
        sheet.cell(row, 3).number_format = "0.0%"

    sheet.column_dimensions["A"].width = 27
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["C"].width = 36
    sheet.freeze_panes = "A3"


def main() -> None:
    args = parse_args()
    workbook_path = args.workbook.resolve()
    audio_dir = args.audio_dir.resolve()
    output_path = (
        args.output.resolve()
        if args.output
        else default_output_path(workbook_path)
    )

    if not workbook_path.exists():
        raise FileNotFoundError(
            f"엑셀 파일을 찾을 수 없습니다: {workbook_path}"
        )
    if not audio_dir.exists():
        raise FileNotFoundError(
            f"음성 폴더를 찾을 수 없습니다: {audio_dir}"
        )

    keep_vba = workbook_path.suffix.lower() == ".xlsm"
    workbook = load_workbook(workbook_path, keep_vba=keep_vba)
    if SHEET_NAME not in workbook.sheetnames:
        raise KeyError(
            f"'{SHEET_NAME}' 시트를 찾을 수 없습니다. "
            f"현재 시트: {', '.join(workbook.sheetnames)}"
        )

    sheet = workbook[SHEET_NAME]
    prepare_columns(sheet)

    tested = 0
    missing_audio = 0
    raw_exact = 0
    normalized_exact = 0
    cer_sum = 0.0
    wer_sum = 0.0
    command_correct = 0
    command_tested = 0
    field_correct: Counter = Counter()
    field_total: Counter = Counter()

    for row in range(2, sheet.max_row + 1):
        number = sheet.cell(row, COL_NUMBER).value
        reference = sheet.cell(row, COL_SENTENCE).value

        if not isinstance(number, (int, float)) or not reference:
            continue

        number = int(number)
        if number < args.start or number > args.end:
            continue

        audio_path = audio_dir / f"{number:03d}.wav"
        sheet.cell(row, COL_AUDIO).value = audio_path.name

        if not audio_path.exists():
            missing_audio += 1
            sheet.cell(row, COL_NOTE).value = "음성 파일 없음"
            print(f"[{number}/100] 음성 파일 없음: {audio_path}")
            continue

        print(f"[{number}/100] Whisper 인식 중")

        try:
            existing = sheet.cell(row, COL_WHISPER).value
            if args.resume and existing:
                whisper_text = str(existing)
            else:
                audio = read_audio(audio_path)
                whisper_text = transcribe_audio(audio)

            normalized_reference = normalize_voice_command(str(reference))
            normalized = normalize_voice_command(whisper_text)
            cer, wer = text_metrics(str(reference), whisper_text)
            _, normalized_wer = text_metrics(
                normalized_reference,
                normalized,
            )

            tested += 1
            cer_sum += cer
            wer_sum += wer
            raw_match = clean_for_cer(str(reference)) == clean_for_cer(
                whisper_text
            )
            normalized_match = clean_for_cer(
                normalized_reference
            ) == clean_for_cer(normalized)
            raw_exact += int(raw_match)
            normalized_exact += int(normalized_match)

            sheet.cell(row, COL_WHISPER).value = whisper_text
            sheet.cell(row, COL_NORMALIZED).value = normalized
            sheet.cell(row, COL_CER).value = cer
            sheet.cell(row, COL_WER).value = wer
            sheet.cell(row, COL_RAW_EXACT).value = (
                "일치" if raw_match else "불일치"
            )
            sheet.cell(row, COL_NORMALIZED_EXACT).value = (
                "일치" if normalized_match else "불일치"
            )

            notes = [f"정규화 후 WER={normalized_wer:.3f}"]

            if not args.skip_llm:
                parsed_commands, raw_json = process(normalized)
                actual, command_errors = commands_to_actual(parsed_commands)
                expected = expected_from_row(sheet, row)
                mismatches, field_results = compare_commands(expected, actual)

                command_tested += 1
                for field, correct in field_results.items():
                    field_total[field] += 1
                    if correct:
                        field_correct[field] += 1

                sheet.cell(row, COL_LLM_JSON).value = raw_json
                if command_errors:
                    sheet.cell(row, COL_COMMAND_RESULT).value = "오류"
                elif mismatches:
                    sheet.cell(row, COL_COMMAND_RESULT).value = "오답"
                else:
                    command_correct += 1
                    sheet.cell(row, COL_COMMAND_RESULT).value = "정답"

                notes.extend(command_errors)
                notes.extend(mismatches)

            sheet.cell(row, COL_NOTE).value = "\n".join(notes)

            for column in (
                COL_WHISPER,
                COL_NORMALIZED,
                COL_LLM_JSON,
                COL_NOTE,
            ):
                sheet.cell(row, column).alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )
            sheet.cell(row, COL_CER).number_format = "0.0%"
            sheet.cell(row, COL_WER).number_format = "0.0%"

        except Exception as error:
            sheet.cell(row, COL_NOTE).value = (
                f"{type(error).__name__}: {error}"
            )
            print(f"  오류: {error}")

        if tested and tested % 5 == 0:
            write_summary(
                workbook,
                tested,
                raw_exact,
                normalized_exact,
                cer_sum,
                wer_sum,
                command_correct,
                command_tested,
                field_correct,
                field_total,
                missing_audio,
            )
            workbook.save(output_path)
            print(f"  중간 저장: {output_path}")

    write_summary(
        workbook,
        tested,
        raw_exact,
        normalized_exact,
        cer_sum,
        wer_sum,
        command_correct,
        command_tested,
        field_correct,
        field_total,
        missing_audio,
    )
    workbook.save(output_path)

    print("\n음성 정확도 평가 완료")
    print(f"결과 파일: {output_path}")
    if tested:
        print(f"평균 CER: {cer_sum / tested:.1%}")
        print(f"평균 WER: {wer_sum / tested:.1%}")
    if command_tested:
        print(
            "최종 명령 정확도: "
            f"{command_correct / command_tested:.1%}"
        )


if __name__ == "__main__":
    main()
