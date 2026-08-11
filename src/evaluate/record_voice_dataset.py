"""엑셀의 테스트 문장 100개를 차례로 보여주고 WAV 파일로 녹음한다.

이 파일을 프로젝트의 src 폴더에 넣고 실행한다.

사용 예:
    python src/record_voice_dataset.py "예제 100.xlsm"
    python src/record_voice_dataset.py "예제 100.xlsm" --start 21
"""

from __future__ import annotations

import argparse
from pathlib import Path

import scipy.io.wavfile as wav
from openpyxl import load_workbook

from voice.recorder import SAMPLE_RATE, record_until_silence


SHEET_NAME = "예시 문장"
COL_NUMBER = 1
COL_SENTENCE = 11


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="엑셀의 테스트 문장을 읽으며 음성 데이터셋을 녹음합니다."
    )
    parser.add_argument("workbook", type=Path, help="테스트 문장이 있는 엑셀 파일")
    parser.add_argument(
        "--audio-dir",
        type=Path,
        default=Path("voice_test_audio"),
        help="WAV 저장 폴더(기본값: voice_test_audio)",
    )
    parser.add_argument("--start", type=int, default=1, help="시작 번호")
    parser.add_argument("--end", type=int, default=100, help="종료 번호")
    return parser.parse_args()


def show_status(message: str) -> None:
    print(f"\n[{message.replace(chr(10), ' ')}]")


def main() -> None:
    args = parse_args()
    workbook_path = args.workbook.resolve()
    audio_dir = args.audio_dir.resolve()

    if not workbook_path.exists():
        raise FileNotFoundError(
            f"엑셀 파일을 찾을 수 없습니다: {workbook_path}"
        )

    workbook = load_workbook(
        workbook_path,
        read_only=True,
        data_only=True,
    )

    if SHEET_NAME not in workbook.sheetnames:
        raise KeyError(
            f"'{SHEET_NAME}' 시트를 찾을 수 없습니다. "
            f"현재 시트: {', '.join(workbook.sheetnames)}"
        )

    sheet = workbook[SHEET_NAME]
    audio_dir.mkdir(parents=True, exist_ok=True)

    print("\n음성 테스트 데이터 녹음을 시작합니다.")
    print("Enter: 녹음 시작 / s: 건너뛰기 / q: 종료")
    print("녹음 후 Enter: 저장 / p: 재생 / r: 다시 녹음\n")

    for row in range(2, sheet.max_row + 1):
        number = sheet.cell(row, COL_NUMBER).value
        sentence = sheet.cell(row, COL_SENTENCE).value

        if not isinstance(number, (int, float)) or not sentence:
            continue

        number = int(number)
        if number < args.start or number > args.end:
            continue

        output_path = audio_dir / f"{number:03d}.wav"

        print("\n" + "=" * 70)
        print(f"[{number}/100]")
        print(sentence)

        if output_path.exists():
            print(f"기존 녹음 있음: {output_path}")
            command = input(
                "Enter=덮어쓰기 / s=건너뛰기 / q=종료: "
            ).strip().lower()
        else:
            command = input(
                "Enter=녹음 시작 / s=건너뛰기 / q=종료: "
            ).strip().lower()

        if command == "q":
            print("녹음을 종료합니다.")
            break
        if command == "s":
            continue

        while True:
            print("\n주변 소음 측정 중에는 말하지 마세요.")
            audio = record_until_silence(show_status)

            if audio is None:
                retry = input(
                    "녹음 실패. Enter=다시 시도 / s=건너뛰기 / q=종료: "
                ).strip().lower()
                if retry == "q":
                    return
                if retry == "s":
                    break
                continue

            duration = len(audio) / SAMPLE_RATE
            print(f"녹음 길이: {duration:.2f}초")

            while True:
                action = input(
                    "Enter=저장 / p=재생 / r=다시 녹음 "
                    "/ s=건너뛰기 / q=종료: "
                ).strip().lower()

                if action == "p":
                    try:
                        import sounddevice as sd

                        sd.play(audio, SAMPLE_RATE)
                        sd.wait()
                    except Exception as error:
                        print(f"재생 오류: {error}")
                    continue

                if action == "r":
                    break

                if action == "s":
                    audio = None
                    break

                if action == "q":
                    return

                wav.write(output_path, SAMPLE_RATE, audio)
                print(f"저장 완료: {output_path}")
                audio = None
                break

            if audio is None:
                break

    print(f"\n녹음 작업 완료: {audio_dir}")


if __name__ == "__main__":
    main()
