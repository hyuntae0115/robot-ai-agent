from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .exceptions import (
    CuttingDataFormatError,
    CuttingDataNotFoundError,
)
from .material_mapper import MaterialDefinition


PROJECT_ROOT = Path(__file__).resolve().parents[2]

DEFAULT_DATA_DIR = (
    PROJECT_ROOT
    / "data"
    / "cutting_conditions"
)


def normalize_header(
    value: object
) -> str:
    return re.sub(
        r"[^0-9a-z]+",
        "",
        str(value).lower(),
    )


class CuttingDataRepository:

    def __init__(
        self,
        data_dir: Path | str = DEFAULT_DATA_DIR,
    ):
        self.data_dir = Path(data_dir)

    def _get_excel_path(
        self,
        material: MaterialDefinition,
    ) -> Path:
        excel_path = (
            self.data_dir
            / material.excel_filename
        )

        if not excel_path.exists():
            raise FileNotFoundError(
                "절삭조건 Excel 파일을 "
                "찾을 수 없습니다:\n"
                f"{excel_path}"
            )

        return excel_path

    @staticmethod
    def _build_column_map(
        headers: tuple[Any, ...],
    ) -> dict[str, int]:
        columns: dict[str, int] = {}

        for index, header in enumerate(
            headers
        ):
            if header is None:
                continue

            normalized = normalize_header(
                header
            )

            columns[normalized] = index

        return columns

    @staticmethod
    def _require_column(
        columns: dict[str, int],
        header_name: str,
        excel_path: Path,
    ) -> int:
        normalized = normalize_header(
            header_name
        )

        if normalized not in columns:
            raise CuttingDataFormatError(
                "Excel 필수 열이 없습니다: "
                f"{header_name}\n"
                f"파일: {excel_path}"
            )

        return columns[normalized]

    @staticmethod
    def _to_float(
        value: Any,
        field_name: str,
        row_number: int,
    ) -> float:
        try:
            return float(value)

        except (TypeError, ValueError) as error:
            raise CuttingDataFormatError(
                f"{field_name} 값이 "
                "숫자가 아닙니다.\n"
                f"행: {row_number}\n"
                f"값: {value!r}"
            ) from error

    def find_exact(
        self,
        material: MaterialDefinition,
        diameter_mm: float,
        depth_mm: float,
    ) -> dict[str, Any]:
        excel_path = self._get_excel_path(
            material
        )

        workbook = load_workbook(
            excel_path,
            read_only=True,
            data_only=True,
        )

        try:
            worksheet = workbook.active

            if worksheet is None:
                raise CuttingDataFormatError(
                    "활성화된 Excel 시트가 없습니다: "
                    f"{excel_path}"
                )

            row_iterator = worksheet.iter_rows(
                values_only=True
            )

            try:
                headers = next(row_iterator)

            except StopIteration as error:
                raise CuttingDataFormatError(
                    "Excel 파일이 비어 있습니다: "
                    f"{excel_path}"
                ) from error

            columns = self._build_column_map(
                headers
            )

            diameter_column = (
                self._require_column(
                    columns,
                    "diameter(mm)",
                    excel_path,
                )
            )

            depth_column = (
                self._require_column(
                    columns,
                    "depth(mm)",
                    excel_path,
                )
            )

            vc_column = self._require_column(
                columns,
                "V_c (m/min)",
                excel_path,
            )

            feed_column = (
                self._require_column(
                    columns,
                    "Feed(mm/rev)",
                    excel_path,
                )
            )

            rpm_column = self._require_column(
                columns,
                "RPM(rev/min)",
                excel_path,
            )

            tool_column = columns.get(
                normalize_header(
                    "Utilized Tool"
                )
            )

            available_diameters: set[
                float
            ] = set()

            available_depths: set[
                float
            ] = set()

            matched_rows: list[
                dict[str, Any]
            ] = []

            for row_number, row in enumerate(
                row_iterator,
                start=2,
            ):
                diameter_value = row[
                    diameter_column
                ]

                depth_value = row[
                    depth_column
                ]

                if (
                    diameter_value is None
                    or depth_value is None
                ):
                    continue

                row_diameter = self._to_float(
                    diameter_value,
                    "diameter",
                    row_number,
                )

                row_depth = self._to_float(
                    depth_value,
                    "depth",
                    row_number,
                )

                available_diameters.add(
                    row_diameter
                )

                diameter_matches = (
                    abs(
                        row_diameter
                        - diameter_mm
                    )
                    < 1e-9
                )

                if diameter_matches:
                    available_depths.add(
                        row_depth
                    )

                depth_matches = (
                    abs(
                        row_depth
                        - depth_mm
                    )
                    < 1e-9
                )

                if not (
                    diameter_matches
                    and depth_matches
                ):
                    continue

                tool: str | None = None

                if tool_column is not None:
                    tool_value = row[
                        tool_column
                    ]

                    if tool_value is not None:
                        tool = str(tool_value)

                matched_rows.append(
                    {
                        "material": (
                            material.display_name
                        ),
                        "diameter_mm": (
                            row_diameter
                        ),
                        "depth_mm": row_depth,
                        "vc_m_min": (
                            self._to_float(
                                row[vc_column],
                                "V_c",
                                row_number,
                            )
                        ),
                        "feed_mm_rev": (
                            self._to_float(
                                row[feed_column],
                                "Feed(mm/rev)",
                                row_number,
                            )
                        ),
                        "rpm": int(
                            round(
                                self._to_float(
                                    row[rpm_column],
                                    "RPM",
                                    row_number,
                                )
                            )
                        ),
                        "tool": tool,
                        "source_file": str(
                            excel_path
                        ),
                        "source_row": (
                            row_number
                        ),
                    }
                )

            if len(matched_rows) > 1:
                raise CuttingDataFormatError(
                    "동일한 소재·직경·깊이 "
                    "조합이 여러 개 있습니다.\n"
                    f"소재: {material.display_name}\n"
                    f"직경: {diameter_mm} mm\n"
                    f"깊이: {depth_mm} mm"
                )

            if len(matched_rows) == 1:
                return matched_rows[0]

            diameter_text = ", ".join(
                f"{value:g}"
                for value in sorted(
                    available_diameters
                )
            )

            if available_depths:
                depth_text = ", ".join(
                    f"{value:g}"
                    for value in sorted(
                        available_depths
                    )
                )

                raise CuttingDataNotFoundError(
                    f"직경 {diameter_mm:g} mm에는 "
                    f"깊이 {depth_mm:g} mm "
                    "데이터가 없습니다.\n"
                    "가능한 깊이(mm): "
                    f"{depth_text}"
                )

            raise CuttingDataNotFoundError(
                f"직경 {diameter_mm:g} mm "
                "데이터가 없습니다.\n"
                "사용 가능한 직경(mm): "
                f"{diameter_text}"
            )

        finally:
            workbook.close()