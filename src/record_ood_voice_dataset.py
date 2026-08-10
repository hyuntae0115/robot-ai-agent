"""OOD 음성 테스트 문장을 차례로 보여주고 WAV 파일로 녹음한다.

프로젝트의 src 폴더에 넣어 실행한다.

사용 예:
    python src/record_ood_voice_dataset.py "음성인식_OOD_테스트문장_200개.xlsx" --test-type meaningless
    python src/record_ood_voice_dataset.py "음성인식_OOD_테스트문장_200개.xlsx" --test-type unsupported
    python src/record_ood_voice_dataset.py "음성인식_OOD_테스트문장_200개.xlsx" --test-type meaningless --start 51
"""

from __future__ import annotations

import argparse
from pathlib import Path

import scipy.io.wavfile as wav
from openpyxl import load_workbook

from voice.recorder import SAMPLE_RATE, record_until_silence


TEST_TYPES = {
    "meaningless": ("무의미 발화", "meaningless"),
    "unsupported": ("미등록 재질", "unsupported_material"),
}
COL_NUMBER = 1
COL_SENTENCE = 11


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="OOD 테스트 문장을 읽으며 음성 데이터셋을 녹음합니다."
    )
    parser.add_argument("workbook", type=Path, help="OOD 테스트 엑셀 파일")
    parser.add_argument(
        "--test-type",
        choices=tuple(TEST_TYPES),
        required=True,
        help="meaningless=무의미 발화, unsupported=미등록 재질",
    )
    parser.add_argument(
        "--audio-root",
        type=Path,
        default=Path("voice_test_audio"),
        help="WAV 저장 폴더의 상위 경로",
    )
    parser.add_argument("--start", type=int, default=1, help="시작 번호")
    parser.add_argument("--end", type=int, default=100, help="종료 번호")
    return parser.parse_args()


def show_status(message: str) -> None:
    print(f"\n[{message.replace(chr(10), ' ')}]")


def main() -> None:
    args = parse_args()
    workbook_path = args.workbook.resolve()
    sheet_name, folder_name = TEST_TYPES[args.test_type]
    audio_dir = (args.audio_root / folder_name).resolve()

    if not workbook_path.exists():
        raise FileNotFoundError(f"엑셀 파일을 찾을 수 없습니다: {workbook_path}")
    if args.start < 1 or args.end > 100 or args.start > args.end:
        raise ValueError("녹음 범위는 1~100이며 start는 end보다 클 수 없습니다.")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise KeyError(
            f"'{sheet_name}' 시트가 없습니다. 현재 시트: {', '.join(workbook.sheetnames)}"
        )

    sheet = workbook[sheet_name]
    audio_dir.mkdir(parents=True, exist_ok=True)

    print(f"\n시험 유형: {sheet_name}")
    print(f"저장 폴더: {audio_dir}")
    print("Enter: 녹음 시작 / s: 건너뛰기 / q: 종료")
    print("녹음 후 Enter: 저장 / p: 재생 / r: 다시 녹음\n")

    for row in range(2, 102):
        number = sheet.cell(row, COL_NUMBER).value
        sentence = sheet.cell(row, COL_SENTENCE).value
        if not isinstance(number, (int, float)) or not sentence:
            continue

        number = int(number)
        if not args.start <= number <= args.end:
            continue

        output_path = audio_dir / f"{number:03d}.wav"
        print("\n" + "=" * 70)
        print(f"[{sheet_name} {number}/100]")
        print(sentence)

        prompt = (
            "Enter=덮어쓰기 / s=건너뛰기 / q=종료: "
            if output_path.exists()
            else "Enter=녹음 시작 / s=건너뛰기 / q=종료: "
        )
        if output_path.exists():
            print(f"기존 녹음 있음: {output_path}")
        command = input(prompt).strip().lower()
        if command == "q":
            print("녹음을 종료합니다.")
            break
        if command == "s":
            continue

        while True:
            print("\n주변 소음 측정 중에는 말하지 마세요.")
            audio = record_until_silence(show_status)
            if audio is None:
                retry = input(
                    "녹음 실패. Enter=재시도 / s=건너뛰기 / q=종료: "
                ).strip().lower()
                if retry == "q":
                    return
                if retry == "s":
                    break
                continue

            print(f"녹음 길이: {len(audio) / SAMPLE_RATE:.2f}초")
            while True:
                action = input(
                    "Enter=저장 / p=재생 / r=다시 녹음 / s=건너뛰기 / q=종료: "
                ).strip().lower()
                if action == "p":
                    try:
                        import sounddevice as sd

                        sd.play(audio, SAMPLE_RATE)
                        sd.wait()
                    except Exception as error:
                        print(f"재생 오류: {error}")
                    continue
                if action == "r":
                    break
                if action == "s":
                    audio = None
                    break
                if action == "q":
                    return

                wav.write(output_path, SAMPLE_RATE, audio)
                print(f"저장 완료: {output_path}")
                audio = None
                break

            if audio is None:
                break

    workbook.close()
    print(f"\n녹음 작업 완료: {audio_dir}")


if __name__ == "__main__":
    main()
